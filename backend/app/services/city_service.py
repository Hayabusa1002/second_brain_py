from __future__ import annotations

import csv
import io
import json
import yaml
from types import SimpleNamespace
from typing import Any
from uuid import UUID

from fastapi import UploadFile
from openpyxl import load_workbook

from app.repositories.city_repository import CityRepository
from app.schemas.bulk_import import ImportError, ImportResult, ImportLogItem


class CityService:
    def __init__(
        self,
        repository: CityRepository,
    ):
        self.repository = repository

    def list_cities(self):
        return self.repository.list()

    def get_city(self, city_id: UUID):
        return self.repository.get_by_id(city_id)

    def create_city(self, data):
        existing = self.repository.get_by_identity(
            name=data.name,
            state=getattr(data, "state", None),
            country=getattr(data, "country", None),
        )
        if existing:
            return existing
        return self.repository.add(data)

    def update_city(self, city_id: UUID, data):
        return self.repository.update(city_id, data)

    def delete_city(self, city_id: UUID) -> bool:
        return self.repository.delete(city_id)

    async def import_cities(self, file: UploadFile, current_user) -> ImportResult:
        filename = (file.filename or "").lower()

        if filename.endswith(".csv"):
            rows = await self._parse_csv(file)
        elif filename.endswith(".json"):
            rows = await self._parse_json(file)
        elif filename.endswith(".yaml") or filename.endswith(".yml"):
            rows = await self._parse_yaml(file)
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            rows = await self._parse_xlsx(file)
        else:
            raise ValueError("Unsupported format. Use CSV, XLSX, JSON or YAML.")

        total = len(rows)
        processed = 0
        imported = 0
        skipped = 0
        warnings = 0
        errors_count = 0
        errors: list[ImportError] = []
        logs: list[ImportLogItem] = []

        seen_cities: set[tuple[str, str | None, str | None]] = set()

        for index, item in enumerate(rows, start=1):
            processed += 1

            row_errors = self._validate_import_item(item, index)
            if row_errors:
                errors.extend(row_errors)
                errors_count += len(row_errors)
                for err in row_errors:
                    logs.append(
                        ImportLogItem(
                            row=err.row,
                            level="error",
                            entity="city",
                            name=item.get("name") or None,
                            message=err.error,
                        )
                    )
                continue

            city_name = item["name"].strip()
            state = (item.get("state") or "").strip() or None
            country = (item.get("country") or "").strip() or None

            city_key = (city_name.lower(), state.lower() if state else None, country.lower() if country else None)

            if city_key in seen_cities:
                msg = f"Duplicate city in file: '{city_name}' (state={state or '-'}, country={country or '-'})."
                error = ImportError(row=index, error=msg)
                errors.append(error)
                errors_count += 1
                warnings += 1
                logs.append(
                    ImportLogItem(
                        row=index,
                        level="warning",
                        entity="city",
                        name=city_name,
                        message=msg,
                    )
                )
                skipped += 1
                continue

            seen_cities.add(city_key)

            existing = self.repository.get_by_identity(
                name=city_name,
                state=state,
                country=country,
            )

            if not existing:
                city = self.repository.add(
                    SimpleNamespace(
                        name=city_name,
                        state=state,
                        country=country,
                    )
                )
                imported += 1
                logs.append(
                    ImportLogItem(
                        row=index,
                        level="info",
                        entity="city",
                        name=city_name,
                        message="City created.",
                    )
                )
            else:
                skipped += 1
                warnings += 1
                logs.append(
                    ImportLogItem(
                        row=index,
                        level="warning",
                        entity="city",
                        name=city_name,
                        message="City already exists. Using existing record.",
                    )
                )

        return ImportResult(
            total=total,
            processed=processed,
            imported=imported,
            skipped=skipped,
            warnings=warnings,
            errors_count=errors_count,
            errors=errors,
            logs=logs,
        )

    async def _parse_csv(self, file: UploadFile) -> list[dict[str, Any]]:
        content = await file.read()
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))

        rows: list[dict[str, Any]] = []
        for row in reader:
            rows.append(
                {
                    "name": (row.get("name") or "").strip(),
                    "state": (row.get("state") or "").strip(),
                    "country": (row.get("country") or "").strip(),
                }
            )
        return rows

    async def _parse_json(self, file: UploadFile) -> list[dict[str, Any]]:
        content = await file.read()
        data = json.loads(content.decode("utf-8"))

        if not isinstance(data, list):
            raise ValueError("JSON import must be a list of cities.")

        return [self._normalize_structured_item(item) for item in data]

    async def _parse_yaml(self, file: UploadFile) -> list[dict[str, Any]]:
        content = await file.read()
        data = yaml.safe_load(content.decode("utf-8"))

        if not isinstance(data, list):
            raise ValueError("YAML import must be a list of cities.")

        return [self._normalize_structured_item(item) for item in data]

    async def _parse_xlsx(self, file: UploadFile) -> list[dict[str, Any]]:
        content = await file.read()
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active

        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []

        headers = [str(h).strip() if h is not None else "" for h in values[0]]
        rows: list[dict[str, Any]] = []

        for raw_row in values[1:]:
            row_dict = dict(zip(headers, raw_row))

            rows.append(
                {
                    "name": str(row_dict.get("name") or "").strip(),
                    "state": str(row_dict.get("state") or "").strip(),
                    "country": str(row_dict.get("country") or "").strip(),
                }
            )

        return rows

    def _normalize_structured_item(self, item: Any) -> dict[str, Any]:
        if not isinstance(item, dict):
            return {"name": "", "state": "", "country": ""}

        return {
            "name": str(item.get("name") or "").strip(),
            "state": str(item.get("state") or "").strip(),
            "country": str(item.get("country") or "").strip(),
        }

    def _validate_import_item(self, item: dict[str, Any], row: int) -> list[ImportError]:
        errors: list[ImportError] = []

        name = item.get("name", "").strip()
        country = item.get("country", "").strip()

        if not name:
            errors.append(ImportError(row=row, error="City name is required."))

        if not country:
            errors.append(ImportError(row=row, error="Country is required."))

        return errors