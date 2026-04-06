from __future__ import annotations

import csv
import io
import json
import yaml
from types import SimpleNamespace
from typing import Any
from uuid import UUID

from fastapi import UploadFile
from openpyxl import load_workbook

from app.repositories.category_repository import CategoryRepository
from app.repositories.subcategory_repository import SubcategoryRepository
from app.schemas.bulk_import import ImportError, ImportResult, ImportLogItem

ALLOWED_CATEGORY_TYPES = {"income", "expense"}


class CategoryService:
    def __init__(
        self,
        repository: CategoryRepository,
        subcategory_repository: SubcategoryRepository,
    ):
        self.repository = repository
        self.subcategory_repository = subcategory_repository

    def list_categories(self):
        return self.repository.list()

    def get_category(self, category_id: UUID):
        return self.repository.get_by_id(category_id)

    def create_category(self, data):
        existing = self.repository.get_by_name_and_type(data.name, data.type)
        if existing:
            return existing
        return self.repository.add(data)

    def update_category(self, category_id: UUID, data):
        category = self.repository.get_by_id(category_id)
        if not category:
            return None

        if getattr(data, "name", None):
            category_type = getattr(data, "type", category.type)
            existing = self.repository.get_by_name_and_type(data.name, category_type)
            if existing and existing.id != category_id:
                return None

        return self.repository.update(category_id, data)

    def delete_category(self, category_id: UUID) -> bool:
        return self.repository.delete(category_id)

    async def import_categories(self, file: UploadFile, current_user) -> ImportResult:
        filename = (file.filename or "").lower()

        if filename.endswith(".csv"):
            rows = await self._parse_csv(file)
        elif filename.endswith(".json"):
            rows = await self._parse_json(file)
        elif filename.endswith(".yaml") or filename.endswith(".yml"):
            rows = await self._parse_yaml(file)
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            rows = await self._parse_xlsx(file)
        else:
            raise ValueError("Unsupported format. Use CSV, XLSX, JSON or YAML.")

        total = len(rows)
        processed = 0
        imported = 0
        skipped = 0
        warnings = 0
        errors_count = 0
        errors: list[ImportError] = []
        logs: list[ImportLogItem] = []

        seen_categories: set[tuple[str, str]] = set()

        for index, item in enumerate(rows, start=1):
            processed += 1

            row_errors = self._validate_import_item(item, index)
            if row_errors:
                errors.extend(row_errors)
                errors_count += len(row_errors)
                for err in row_errors:
                    logs.append(
                        ImportLogItem(
                            row=err.row,
                            level="error",
                            entity="category",
                            name=item.get("name") or None,
                            message=err.error,
                        )
                    )
                continue

            category_name = item["name"].strip()
            category_type = item["type"].strip().lower()
            category_key = (category_name.lower(), category_type)

            if category_key in seen_categories:
                msg = f"Duplicate category in file: '{category_name}' ({category_type})."
                error = ImportError(row=index, error=msg)
                errors.append(error)
                errors_count += 1
                warnings += 1
                logs.append(
                    ImportLogItem(
                        row=index,
                        level="warning",
                        entity="category",
                        name=category_name,
                        message=msg,
                    )
                )
                skipped += 1
                continue

            seen_categories.add(category_key)

            category = self.repository.get_by_name_and_type(category_name, category_type)
            if not category:
                category = self.repository.add(
                    SimpleNamespace(
                        name=category_name,
                        type=category_type,
                    )
                )
                imported += 1
                logs.append(
                    ImportLogItem(
                        row=index,
                        level="info",
                        entity="category",
                        name=category_name,
                        message="Category created.",
                    )
                )
            else:
                skipped += 1
                warnings += 1
                logs.append(
                    ImportLogItem(
                        row=index,
                        level="warning",
                        entity="category",
                        name=category_name,
                        message="Category already exists. Using existing record.",
                    )
                )

            sub_seen: set[str] = set()
            for sub_name in item.get("subcategories", []):
                normalized_sub_name = sub_name.strip()

                if not normalized_sub_name:
                    msg = f"Empty subcategory name in category '{category_name}'."
                    error = ImportError(row=index, error=msg)
                    errors.append(error)
                    errors_count += 1
                    logs.append(
                        ImportLogItem(
                            row=index,
                            level="error",
                            entity="subcategory",
                            name=None,
                            message=msg,
                        )
                    )
                    continue

                sub_key = normalized_sub_name.lower()
                if sub_key in sub_seen:
                    msg = f"Duplicate subcategory '{normalized_sub_name}' in category '{category_name}'."
                    error = ImportError(row=index, error=msg)
                    errors.append(error)
                    errors_count += 1
                    warnings += 1
                    logs.append(
                        ImportLogItem(
                            row=index,
                            level="warning",
                            entity="subcategory",
                            name=normalized_sub_name,
                            message=msg,
                        )
                    )
                    continue

                sub_seen.add(sub_key)

                existing_sub = self.subcategory_repository.get_by_name_and_category(
                    normalized_sub_name,
                    category.id,
                )
                if existing_sub:
                    warnings += 1
                    logs.append(
                        ImportLogItem(
                            row=index,
                            level="warning",
                            entity="subcategory",
                            name=normalized_sub_name,
                            message="Subcategory already exists. Using existing record.",
                        )
                    )
                    continue

                self.subcategory_repository.add(
                    category_id=category.id,
                    data=SimpleNamespace(name=normalized_sub_name),
                )
                logs.append(
                    ImportLogItem(
                        row=index,
                        level="info",
                        entity="subcategory",
                        name=normalized_sub_name,
                        message="Subcategory created.",
                    )
                )

        return ImportResult(
            total=total,
            processed=processed,
            imported=imported,
            skipped=skipped,
            warnings=warnings,
            errors_count=errors_count,
            errors=errors,
            logs=logs,
        )

    async def _parse_csv(self, file: UploadFile) -> list[dict[str, Any]]:
        content = await file.read()
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))

        rows: list[dict[str, Any]] = []
        for row in reader:
            subcategories = self._split_subcategory_names(row.get("subcategory_names", ""))
            rows.append(
                {
                    "name": (row.get("name") or "").strip(),
                    "type": (row.get("type") or "").strip().lower(),
                    "subcategories": subcategories,
                }
            )
        return rows

    async def _parse_json(self, file: UploadFile) -> list[dict[str, Any]]:
        content = await file.read()
        data = json.loads(content.decode("utf-8"))

        if not isinstance(data, list):
            raise ValueError("JSON import must be a list of categories.")

        return [self._normalize_structured_item(item) for item in data]

    async def _parse_yaml(self, file: UploadFile) -> list[dict[str, Any]]:
        content = await file.read()
        data = yaml.safe_load(content.decode("utf-8"))

        if not isinstance(data, list):
            raise ValueError("YAML import must be a list of categories.")

        return [self._normalize_structured_item(item) for item in data]

    async def _parse_xlsx(self, file: UploadFile) -> list[dict[str, Any]]:
        content = await file.read()
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active

        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []

        headers = [str(h).strip() if h is not None else "" for h in values[0]]
        rows: list[dict[str, Any]] = []

        for raw_row in values[1:]:
            row_dict = dict(zip(headers, raw_row))
            subcategories = self._split_subcategory_names(row_dict.get("subcategory_names", ""))

            rows.append(
                {
                    "name": str(row_dict.get("name") or "").strip(),
                    "type": str(row_dict.get("type") or "").strip().lower(),
                    "subcategories": subcategories,
                }
            )

        return rows

    def _normalize_structured_item(self, item: Any) -> dict[str, Any]:
        if not isinstance(item, dict):
            return {"name": "", "type": "", "subcategories": []}

        raw_subcategories = item.get("subcategories", [])
        normalized_subcategories: list[str] = []

        if isinstance(raw_subcategories, list):
            for sub in raw_subcategories:
                if isinstance(sub, dict):
                    normalized_subcategories.append(str(sub.get("name") or "").strip())
                else:
                    normalized_subcategories.append(str(sub).strip())

        return {
            "name": str(item.get("name") or "").strip(),
            "type": str(item.get("type") or "").strip().lower(),
            "subcategories": normalized_subcategories,
        }

    def _split_subcategory_names(self, value: Any) -> list[str]:
        if value is None:
            return []

        text = str(value).strip()
        if not text:
            return []

        return [part.strip() for part in text.split("|") if part and part.strip()]

    def _validate_import_item(self, item: dict[str, Any], row: int) -> list[ImportError]:
        errors: list[ImportError] = []

        name = item.get("name", "").strip()
        category_type = item.get("type", "").strip().lower()
        subcategories = item.get("subcategories", [])

        if not name:
            errors.append(ImportError(row=row, error="Category name is required."))

        if not category_type:
            errors.append(ImportError(row=row, error="Category type is required."))
        elif category_type not in ALLOWED_CATEGORY_TYPES:
            errors.append(
                ImportError(
                    row=row,
                    error="Category type must be 'income' or 'expense'.",
                )
            )

        if not isinstance(subcategories, list):
            errors.append(
                ImportError(row=row, error="Subcategories must be a list.")
            )

        return errors