import csv
import io
import json
import yaml

from typing import Any
from fastapi import UploadFile
from openpyxl import load_workbook

from app.schemas.bulk_import import ImportError, ImportLogItem, ImportResult


class UnsupportedImportFormatError(Exception):
    def __init__(self, message: str = "Unsupported format. Use CSV, XLSX, JSON or YAML."):
        super().__init__(message)


class BulkImportCollector:
    def __init__(self, total: int = 0):
        self.total = total
        self.processed = 0
        self.imported = 0
        self.skipped = 0
        self.warnings = 0
        self.errors_count = 0
        self.errors: list[ImportError] = []
        self.logs: list[ImportLogItem] = []

    def add_info(
        self,
        row: int,
        entity: str,
        message: str,
        name: str | None = None,
    ) -> None:
        self.logs.append(
            ImportLogItem(
                row=row,
                level="info",
                entity=entity,
                name=name,
                message=message,
            )
        )

    def add_warning(
        self,
        row: int,
        entity: str,
        message: str,
        name: str | None = None,
        count_as_skip: bool = False,
        count_as_error: bool = False,
    ) -> None:
        self.warnings += 1

        if count_as_skip:
            self.skipped += 1

        if count_as_error:
            self.errors.append(ImportError(row=row, error=message))
            self.errors_count += 1

        self.logs.append(
            ImportLogItem(
                row=row,
                level="warning",
                entity=entity,
                name=name,
                message=message,
            )
        )

    def add_error(
        self,
        row: int,
        entity: str,
        message: str,
        name: str | None = None,
    ) -> None:
        self.errors.append(ImportError(row=row, error=message))
        self.errors_count += 1
        self.logs.append(
            ImportLogItem(
                row=row,
                level="error",
                entity=entity,
                name=name,
                message=message,
            )
        )

    def add_row_errors(
        self,
        row_errors: list[ImportError],
        entity: str,
        name: str | None = None,
    ) -> None:
        self.errors.extend(row_errors)
        self.errors_count += len(row_errors)

        for err in row_errors:
            self.logs.append(
                ImportLogItem(
                    row=err.row,
                    level="error",
                    entity=entity,
                    name=name,
                    message=err.error,
                )
            )

    def build(self) -> ImportResult:
        return ImportResult(
            total=self.total,
            processed=self.processed,
            imported=self.imported,
            skipped=self.skipped,
            warnings=self.warnings,
            errors_count=self.errors_count,
            errors=self.errors,
            logs=self.logs,
        )


class BulkImportService:
    async def parse_file(self, file: UploadFile) -> list[dict[str, Any]]:
        filename = (file.filename or "").lower()

        parsers = {
            ".csv": self._parse_csv,
            ".json": self._parse_json,
            ".yaml": self._parse_yaml,
            ".yml": self._parse_yaml,
            ".xlsx": self._parse_xlsx,
            ".xls": self._parse_xlsx,
        }

        for extension, parser in parsers.items():
            if filename.endswith(extension):
                return await parser(file)

        raise UnsupportedImportFormatError()

    async def _parse_csv(self, file: UploadFile) -> list[dict[str, Any]]:
        content = await file.read()
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(row) for row in reader]

    async def _parse_json(self, file: UploadFile) -> list[dict[str, Any]]:
        content = await file.read()
        data = json.loads(content.decode("utf-8"))

        if not isinstance(data, list):
            raise ValueError("JSON import must be a list of records.")

        return [item if isinstance(item, dict) else {} for item in data]

    async def _parse_yaml(self, file: UploadFile) -> list[dict[str, Any]]:
        content = await file.read()
        data = yaml.safe_load(content.decode("utf-8"))

        if not isinstance(data, list):
            raise ValueError("YAML import must be a list of records.")

        return [item if isinstance(item, dict) else {} for item in data]

    async def _parse_xlsx(self, file: UploadFile) -> list[dict[str, Any]]:
        content = await file.read()
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active

        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []

        headers = [str(h).strip() if h is not None else "" for h in values[0]]
        rows: list[dict[str, Any]] = []

        for raw_row in values[1:]:
            row_dict = dict(zip(headers, raw_row))
            rows.append(
                {
                    key: value
                    for key, value in row_dict.items()
                }
            )

        return rows