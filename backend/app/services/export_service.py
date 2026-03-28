from typing import List
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
import json
import csv
from app.models.transaction import Transaction


def _rows(transactions: List[Transaction]) -> List[dict]:
    """Normalize fields"""
    return [
        {
            "id":          str(t.id),
            "date":        str(t.date),
            "type":        t.type.value,
            "amount":      str(t.amount),
            "account":     t.account.name  if t.account  else str(t.account_id),
            "category":    t.category.name if t.category else str(t.category_id),
            "description": t.description or "",
        }
        for t in transactions
    ]


HEADERS = ["id", "date", "type", "amount", "account", "category", "description"]
HEADERS_DISPLAY = ["ID", "Date", "Type", "Amount", "Account", "Category", "Description"]


class ExportService:

    def to_json(self, transactions: List[Transaction]) -> bytes:
        return json.dumps(_rows(transactions), indent=2, ensure_ascii=False).encode("utf-8")

    def to_csv(self, transactions: List[Transaction]) -> bytes:
        buf = BytesIO()
        wrapper = buf  # csv necesita texto, usamos StringIO
        import io
        text_buf = io.StringIO()
        writer = csv.DictWriter(text_buf, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(_rows(transactions))
        return text_buf.getvalue().encode("utf-8")

    def to_xlsx(self, transactions: List[Transaction]) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        ws.append(HEADERS_DISPLAY)

        # Header style
        for cell in ws[1]:
            cell.font      = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill      = openpyxl.styles.PatternFill("solid", fgColor="1a56db")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        for row in _rows(transactions):
            ws.append([
                row["id"],
                row["date"],
                row["type"],
                float(row["amount"]),
                row["account"],
                row["category"],
                row["description"],
            ])

        # Auto column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def to_pdf(self, transactions: List[Transaction]) -> bytes:
        buf    = BytesIO()
        doc    = SimpleDocTemplate(buf, pagesize=A4, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        story  = []

        story.append(Paragraph("Transactions Report", styles["Title"]))
        story.append(Spacer(1, 12))

        pdf_headers = ["Date", "Type", "Amount", "Account", "Category", "Description"]
        rows = [pdf_headers] + [
            [
                row["date"],
                row["type"],
                f"${float(row['amount']):,.2f}",
                row["account"],
                row["category"],
                row["description"][:40],
            ]
            for row in _rows(transactions)
        ]

        table = Table(rows, repeatRows=1, hAlign="LEFT")
        table.setStyle(TableStyle([
            # Header
            ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#1a56db")),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0), 9),
            ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
            # Body
            ("FONTSIZE",      (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ]))

        story.append(table)
        doc.build(story)
        buf.seek(0)
        return buf.read()